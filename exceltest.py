#!/usr/bin/env python3
import openpyxl
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

#wb = openpyxl.load_workbook('sample.xlsx')
wb = load_workbook(filename = 'sample.xlsx')
ws = wb.get_sheet_by_name('Names')
mylista = []
nextList = []
#mydict = ()
# for row in ws.iter_rows('A{}:A{}'.format(ws.min_row,ws.max_row)):
for row in ws.iter_rows('A{}:A{}'):
    for cell in row:
        mylista.append(cell.value)
print(mylista)
print("\n\n\n")

sheet_ranges = wb['Names']
print(sheet_ranges['A1'].value)


# This returns every cell in the worksheet
#for row in ws.iter_rows():
#    for cell in row:
#        nextList.append(cell.value)
#print(nextList)

# fill a blank dictionary through iteration?
